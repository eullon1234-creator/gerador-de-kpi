import warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

warnings.filterwarnings("ignore")

MESES_PT = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR",
    5: "MAI", 6: "JUN", 7: "JUL", 8: "AGO",
    9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}

# ── Paleta de Cores ───────────────────────────────────────────────────────────
C_AZUL_ESCURO  = "1F3864"
C_AZUL_MEDIO   = "2E74B5"
C_AZUL_CLARO   = "BDD7EE"
C_LARANJA      = "ED7D31"
C_LARANJA_LEVE = "FCE4D6"
C_VERDE        = "375623"
C_VERDE_CLARO  = "E2EFDA"
C_BRANCO       = "FFFFFF"
C_VERMELHO     = "C00000"
C_VERMELHO_BG  = "FFE7E7"
C_CINZA        = "F2F2F2"
C_PRETO        = "000000"

# ── Helpers de Estilo ─────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_PRETO, size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Utilitários ───────────────────────────────────────────────────────────────
def _achar_col(colunas, *termos):
    """Encontra coluna: tenta match exato, depois startswith, depois contains."""
    colunas = list(colunas)
    for term in termos:
        for col in colunas:
            if str(col).upper().strip() == term.upper().strip():
                return col
    for term in termos:
        for col in colunas:
            if str(col).upper().strip().startswith(term.upper().strip()):
                return col
    for term in termos:
        for col in colunas:
            if term.upper().strip() in str(col).upper().strip():
                return col
    return None


def _ler_aba(caminho, *nomes):
    """Lê aba com fallback case-insensitive. Retorna DataFrame ou lança ValueError."""
    import openpyxl as _xl
    wb = _xl.load_workbook(caminho, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    nomes_upper = [n.strip().upper() for n in nomes]

    matched = None
    for sheet in sheets:
        if sheet.strip().upper() in nomes_upper:
            matched = sheet
            break

    if not matched:
        raise ValueError(
            f"Aba '{nomes[0]}' não encontrada. "
            f"Abas disponíveis na planilha: {sheets}"
        )

    df = pd.read_excel(caminho, sheet_name=matched, header=0)
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.upper()
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
    )
    return df


def _val(row, col):
    """Lê valor de uma linha de DataFrame com fallback seguro."""
    if col is None:
        return ""
    return row[col] if col in row.index else ""


# ── Função Principal ──────────────────────────────────────────────────────────
def gerar_kpi(caminho_entrada, caminho_saida, data_inicio=None, data_fim=None, meses_morto=3, limite_abc_a=0.8, limite_abc_b=0.95):

    # 1. Ler abas
    df_saida = _ler_aba(caminho_entrada, "SAIDA", "SAÍDA", "Saida", "Saída")
    df_estoque = _ler_aba(caminho_entrada, "ESTOQUE", "Estoque", "estoque")

    # 2. Mapear colunas – SAÍDA
    s_data  = _achar_col(df_saida.columns, "DATA")
    s_cod   = _achar_col(df_saida.columns, "COD")
    s_grupo = _achar_col(df_saida.columns, "GRUPO")
    s_desc  = _achar_col(df_saida.columns, "DESCRI")
    s_unid  = _achar_col(df_saida.columns, "UNID", "UN")
    s_quant = _achar_col(df_saida.columns, "QUANT")

    for nome, val in [("DATA", s_data), ("COD", s_cod), ("GRUPO", s_grupo),
                      ("DESCRIÇÃO", s_desc), ("QUANT", s_quant)]:
        if val is None:
            raise ValueError(
                f"Coluna '{nome}' não encontrada na aba SAIDA. "
                f"Colunas encontradas: {list(df_saida.columns)}"
            )

    # 3. Mapear colunas – ESTOQUE
    e_cod    = _achar_col(df_estoque.columns, "COD")
    e_grupo  = _achar_col(df_estoque.columns, "GRUPO")
    e_desc   = _achar_col(df_estoque.columns, "DESCRI")
    e_saldo  = _achar_col(df_estoque.columns, "SALDO")
    e_vunit  = _achar_col(df_estoque.columns, "VALOR UNIT", "VUNIT", "UNIT")
    e_vtotal = _achar_col(df_estoque.columns, "VALOR TOTAL", "VTOTAL")
    e_unid_e = _achar_col(df_estoque.columns, "UNID", "UN")

    # 4. Limpar SAÍDA
    df_saida[s_data]  = pd.to_datetime(df_saida[s_data], dayfirst=True, errors="coerce")
    df_saida[s_quant] = pd.to_numeric(df_saida[s_quant], errors="coerce").fillna(0)
    df_saida = df_saida.dropna(subset=[s_data])
    df_saida = df_saida[df_saida[s_quant] > 0].copy()

    # 4b. Filtro de período
    if data_inicio:
        df_saida = df_saida[df_saida[s_data] >= pd.Timestamp(data_inicio)]
    if data_fim:
        df_saida = df_saida[df_saida[s_data] <= pd.Timestamp(data_fim)]

    if df_saida.empty:
        raise ValueError(
            "Nenhum registro de saída encontrado no período selecionado. "
            "Verifique as datas e tente novamente."
        )

    df_saida["_PERIODO"] = df_saida[s_data].dt.to_period("M")

    # 5. Limpar ESTOQUE
    for col in [e_saldo, e_vunit, e_vtotal]:
        if col:
            df_estoque[col] = pd.to_numeric(df_estoque[col], errors="coerce").fillna(0)
    if e_cod:
        df_estoque = df_estoque.dropna(subset=[e_cod]).copy()

    # 6. Criar Workbook
    wb = Workbook()
    wb.remove(wb.active)

    periodos = sorted(df_saida["_PERIODO"].unique())
    resumo_dados = []

    for periodo in periodos:
        df_mes = df_saida[df_saida["_PERIODO"] == periodo].copy()
        mes_str = f"{MESES_PT[periodo.month]}-{periodo.year}"

        total_saidas    = int(df_mes[s_quant].sum())
        itens_distintos = int(df_mes[s_cod].nunique())
        grupo_lider     = df_mes.groupby(s_grupo)[s_quant].sum().idxmax()

        # Top 20
        group_cols = [c for c in [s_cod, s_grupo, s_desc, s_unid] if c is not None]
        top20 = (
            df_mes.groupby(group_cols, as_index=False)
            .agg({s_quant: "sum"})
            .sort_values(s_quant, ascending=False)
            .head(20)
            .reset_index(drop=True)
        )
        top20["_PCT"] = top20[s_quant] / total_saidas if total_saidas > 0 else 0

        # Por grupo
        por_grupo = (
            df_mes.groupby(s_grupo)
            .agg(QTD_TOTAL=(s_quant, "sum"), N_ITENS=(s_cod, "nunique"))
            .reset_index()
            .sort_values("QTD_TOTAL", ascending=False)
        )
        por_grupo["_PCT"] = por_grupo["QTD_TOTAL"] / total_saidas if total_saidas > 0 else 0

        resumo_dados.append({
            "mes": mes_str,
            "total_saidas": total_saidas,
            "itens_distintos": itens_distintos,
            "grupo_lider": grupo_lider,
        })

        ws = wb.create_sheet(mes_str)
        _gerar_aba_mes(ws, mes_str, total_saidas, itens_distintos, grupo_lider,
                       top20, por_grupo, s_cod, s_grupo, s_desc, s_unid, s_quant)

    # 7. Resumo Geral (primeira aba)
    ws_resumo = wb.create_sheet("RESUMO GERAL", 0)
    _gerar_resumo_geral(ws_resumo, resumo_dados)

    # 8. Classificação ABC
    ws_abc = wb.create_sheet("CLASS. ABC")
    _gerar_abc(ws_abc, df_saida, s_cod, s_grupo, s_desc, s_unid, s_quant, limite_abc_a, limite_abc_b)

    # 9. Valor por Categoria (requer ESTOQUE com valores)
    if e_grupo and e_cod and (e_vtotal or (e_saldo and e_vunit)):
        ws_cat = wb.create_sheet("VALOR POR CATEGORIA")
        _gerar_valor_categoria(ws_cat, df_estoque, e_cod, e_grupo, e_saldo, e_vunit, e_vtotal)

    # 10. Estoque Morto
    if e_saldo and e_cod:
        ws_morto = wb.create_sheet("ESTOQUE MORTO")
        _gerar_estoque_morto(ws_morto, df_estoque, df_saida, s_cod, s_data,
                             e_cod, e_grupo, e_desc, e_saldo, e_unid_e, e_vunit, meses_morto)

    # 11. Alerta de Estoque (formatação condicional)
    if e_saldo and e_cod:
        ws_alerta = wb.create_sheet("⚠ ALERTA ESTOQUE")
        _gerar_alerta_estoque(ws_alerta, df_estoque, df_saida, s_cod, s_quant, s_data,
                              e_cod, e_grupo, e_desc, e_saldo, e_unid_e, e_vunit)

    wb.save(caminho_saida)


# ── Aba Mensal ────────────────────────────────────────────────────────────────
def _gerar_aba_mes(ws, mes_str, total_saidas, itens_distintos, grupo_lider,
                   top20, por_grupo, s_cod, s_grupo, s_desc, s_unid, s_quant):

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 11

    ws.row_dimensions[1].height = 6

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"KPI DE SAÍDA DE MATERIAIS  |  {mes_str}"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=14)
    c.alignment = _align()
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 10

    # ── Cards – headers ──────────────────────────────────────────────────────
    ws.merge_cells("A5:B5")
    ws["A5"].value = "TOTAL SAÍDAS"
    ws["A5"].fill = _fill(C_AZUL_MEDIO)
    ws["A5"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws["A5"].alignment = _align()

    ws.merge_cells("C5:D5")
    ws["C5"].value = "ITENS DISTINTOS"
    ws["C5"].fill = _fill(C_AZUL_MEDIO)
    ws["C5"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws["C5"].alignment = _align()

    ws.merge_cells("E5:G5")
    ws["E5"].value = "GRUPO LÍDER"
    ws["E5"].fill = _fill(C_LARANJA)
    ws["E5"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws["E5"].alignment = _align()

    ws.row_dimensions[5].height = 18

    # ── Cards – valores ──────────────────────────────────────────────────────
    ws.merge_cells("A6:B6")
    ws["A6"].value = total_saidas
    ws["A6"].fill = _fill(C_AZUL_CLARO)
    ws["A6"].font = _font(bold=True, size=22)
    ws["A6"].alignment = _align()

    ws.merge_cells("C6:D6")
    ws["C6"].value = itens_distintos
    ws["C6"].fill = _fill(C_AZUL_CLARO)
    ws["C6"].font = _font(bold=True, size=22)
    ws["C6"].alignment = _align()

    ws.merge_cells("E6:G6")
    ws["E6"].value = grupo_lider
    ws["E6"].fill = _fill(C_LARANJA_LEVE)
    ws["E6"].font = _font(bold=True, color=C_LARANJA, size=16)
    ws["E6"].alignment = _align()

    ws.row_dimensions[6].height = 42
    ws.row_dimensions[7].height = 10

    # ── Título Top 20 ────────────────────────────────────────────────────────
    ws.merge_cells("A9:G9")
    c = ws["A9"]
    c.value = f"TOP 20 MATERIAIS MAIS SAÍDOS  -  {mes_str}"
    c.fill = _fill(C_AZUL_MEDIO)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()
    ws.row_dimensions[9].height = 20

    # ── Cabeçalho tabela ─────────────────────────────────────────────────────
    COLS = ["A", "B", "C", "D", "E", "F", "G"]
    headers = ["#", "COD", "GRUPO", "DESCRICAO DO MATERIAL", "UN", "QTD SAÍDA", "% DO TOTAL"]
    for col, hdr in zip(COLS, headers):
        c = ws[f"{col}10"]
        c.value = hdr
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[10].height = 18

    # ── Dados Top 20 ─────────────────────────────────────────────────────────
    row = 11
    for i, (_, linha) in enumerate(top20.iterrows(), start=1):
        bg = C_CINZA if i % 2 == 0 else C_BRANCO
        eh_top = i == 1
        dados = [
            i,
            _val(linha, s_cod),
            _val(linha, s_grupo),
            _val(linha, s_desc),
            _val(linha, s_unid) if s_unid else "",
            int(linha[s_quant]),
            linha["_PCT"],
        ]
        alinhamentos = ["center", "center", "center", "left", "center", "center", "center"]

        for j, (col, val, al) in enumerate(zip(COLS, dados, alinhamentos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _align(h=al, wrap=(j == 3))
            if j == 6:
                c.number_format = "0.0%"
                c.font = _font(size=10)
            elif j == 5 and eh_top:
                c.font = _font(bold=True, color=C_VERMELHO, size=11)
            else:
                c.font = _font(bold=eh_top, size=10)

        ws.row_dimensions[row].height = 15
        row += 1

    row += 1  # linha vazia

    # ── Saída por Grupo ──────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = "SAÍDA POR GRUPO"
    c.fill = _fill(C_LARANJA)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()
    ws.row_dimensions[row].height = 20
    row += 1

    hdrs_g = ["GRUPO", "QTD TOTAL", "PARTICIPAÇÃO %", "Nº ITENS"]
    COLS_G = ["A", "B", "C", "D"]
    for col, hdr in zip(COLS_G, hdrs_g):
        c = ws[f"{col}{row}"]
        c.value = hdr
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[row].height = 18
    row += 1

    for i, (_, g) in enumerate(por_grupo.iterrows()):
        bg = C_CINZA if i % 2 == 0 else C_BRANCO
        vals = [g[s_grupo], int(g["QTD_TOTAL"]), g["_PCT"], int(g["N_ITENS"])]
        for col, val in zip(COLS_G, vals):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = _align()
            c.border = _border()
        ws[f"C{row}"].number_format = "0.0%"
        ws.row_dimensions[row].height = 15
        row += 1


# ── Resumo Geral ──────────────────────────────────────────────────────────────
def _gerar_resumo_geral(ws, resumo_dados):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22

    ws.row_dimensions[1].height = 6

    ws.merge_cells("A2:D2")
    ws["A2"].value = "RESUMO GERAL — KPI DE SAÍDA DE MATERIAIS"
    ws["A2"].fill = _fill(C_AZUL_ESCURO)
    ws["A2"].font = _font(bold=True, color=C_BRANCO, size=14)
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 10

    hdrs = ["MÊS", "TOTAL SAÍDAS", "ITENS DISTINTOS", "GRUPO LÍDER"]
    COLS = ["A", "B", "C", "D"]
    for col, hdr in zip(COLS, hdrs):
        c = ws[f"{col}4"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[4].height = 18

    row = 5
    total_geral = 0
    for i, d in enumerate(resumo_dados):
        bg = C_CINZA if i % 2 == 0 else C_BRANCO
        vals = [d["mes"], d["total_saidas"], d["itens_distintos"], d["grupo_lider"]]
        for col, val in zip(COLS, vals):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = _align()
            c.border = _border()
        total_geral += d["total_saidas"]
        ws.row_dimensions[row].height = 15
        row += 1

    # Totalizador
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = "TOTAL GERAL"
    ws[f"A{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"A{row}"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws[f"A{row}"].alignment = _align()
    ws[f"D{row}"].value = total_geral
    ws[f"D{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"D{row}"].font = _font(bold=True, color=C_BRANCO, size=12)
    ws[f"D{row}"].alignment = _align()
    ws.row_dimensions[row].height = 20


# ── Classificação ABC ─────────────────────────────────────────────────────────
def _gerar_abc(ws, df_saida, s_cod, s_grupo, s_desc, s_unid, s_quant, limite_abc_a, limite_abc_b):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 7
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10

    ws.row_dimensions[1].height = 6

    ws.merge_cells("A2:H2")
    ws["A2"].value = "CLASSIFICAÇÃO ABC — MATERIAIS POR VOLUME DE SAÍDA"
    ws["A2"].fill = _fill(C_AZUL_ESCURO)
    ws["A2"].font = _font(bold=True, color=C_BRANCO, size=13)
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 10

    # Legendas
    for col, classe, cor_bg, desc in [
        ("A", "CLASSE A", "375623", f"Representa {int(limite_abc_a*100)}% do volume de saídas"),
        ("D", "CLASSE B", "833C00", f"Representa {int((limite_abc_b-limite_abc_a)*100)}% do volume de saídas"),
        ("F", "CLASSE C", C_VERMELHO, f"Representa {int((1-limite_abc_b)*100)}% do volume de saídas"),
    ]:
        ws[f"{col}4"].value = classe
        ws[f"{col}4"].fill = _fill(cor_bg)
        ws[f"{col}4"].font = _font(bold=True, color=C_BRANCO, size=10)
        ws[f"{col}4"].alignment = _align()
        next_col = chr(ord(col) + 1)
        ws.merge_cells(f"{next_col}4:{next_col}4")
        ws[f"{next_col}4"].value = desc
        ws[f"{next_col}4"].fill = _fill(C_CINZA)
        ws[f"{next_col}4"].font = _font(size=9)
        ws[f"{next_col}4"].alignment = _align(h="left")

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 10

    hdrs = ["#", "COD", "GRUPO", "DESCRIÇÃO", "UN", "QTD TOTAL", "% ACUM.", "CLASSE"]
    COLS = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, hdr in zip(COLS, hdrs):
        c = ws[f"{col}6"]
        c.value = hdr
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[6].height = 18

    group_cols = [c for c in [s_cod, s_grupo, s_desc, s_unid] if c is not None]
    abc = (
        df_saida.groupby(group_cols, as_index=False)
        .agg({s_quant: "sum"})
        .sort_values(s_quant, ascending=False)
        .reset_index(drop=True)
    )
    total = abc[s_quant].sum()
    abc["_PCT"]   = abc[s_quant] / total
    abc["_ACUM"]  = abc["_PCT"].cumsum()
    abc["_CLASSE"] = abc["_ACUM"].apply(lambda x: "A" if x <= limite_abc_a else ("B" if x <= limite_abc_b else "C"))

    COR_CLASSE = {
        "A": ("375623", "E2EFDA"),
        "B": ("833C00", "FCE4D6"),
        "C": (C_VERMELHO, C_VERMELHO_BG),
    }

    row = 7
    for i, (_, linha) in enumerate(abc.iterrows(), start=1):
        classe = linha["_CLASSE"]
        fg, bg = COR_CLASSE[classe]
        dados = [
            i,
            _val(linha, s_cod),
            _val(linha, s_grupo),
            _val(linha, s_desc),
            _val(linha, s_unid) if s_unid else "",
            int(linha[s_quant]),
            linha["_ACUM"],
            classe,
        ]
        alinhamentos = ["center", "center", "center", "left", "center", "center", "center", "center"]
        for j, (col, val, al) in enumerate(zip(COLS, dados, alinhamentos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _align(h=al, wrap=(j == 3))
            if j == 6:
                c.number_format = "0.0%"
                c.font = _font(size=9)
            elif j == 7:
                c.fill = _fill(fg)
                c.font = _font(bold=True, color=C_BRANCO, size=10)
            else:
                c.font = _font(size=9)
        ws.row_dimensions[row].height = 13
        row += 1


# ── Valor por Categoria ───────────────────────────────────────────────────────
def _gerar_valor_categoria(ws, df_estoque, e_cod, e_grupo, e_saldo, e_vunit, e_vtotal):
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    ws.row_dimensions[1].height = 6

    ws.merge_cells("A2:D2")
    ws["A2"].value = "VALOR TOTAL DO ESTOQUE POR CATEGORIA"
    ws["A2"].fill = _fill(C_AZUL_ESCURO)
    ws["A2"].font = _font(bold=True, color=C_BRANCO, size=13)
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 10

    df = df_estoque.copy()
    if e_vtotal and df[e_vtotal].sum() > 0:
        df["_VT"] = df[e_vtotal]
    elif e_saldo and e_vunit:
        df["_VT"] = df[e_saldo] * df[e_vunit]
    else:
        ws.merge_cells("A4:D4")
        ws["A4"].value = "Dados insuficientes para calcular valor por categoria."
        return

    cat = (
        df.groupby(e_grupo)
        .agg(VALOR_TOTAL=("_VT", "sum"), N_ITENS=(e_cod, "nunique"))
        .reset_index()
        .sort_values("VALOR_TOTAL", ascending=False)
    )
    total_val = cat["VALOR_TOTAL"].sum()
    cat["_PCT"] = cat["VALOR_TOTAL"] / total_val if total_val > 0 else 0

    hdrs = ["GRUPO / CATEGORIA", "VALOR TOTAL (R$)", "% DO TOTAL", "Nº ITENS"]
    COLS = ["A", "B", "C", "D"]
    for col, hdr in zip(COLS, hdrs):
        c = ws[f"{col}4"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[4].height = 18

    row = 5
    for i, (_, g) in enumerate(cat.iterrows()):
        bg = C_CINZA if i % 2 == 0 else C_BRANCO
        vals = [g[e_grupo], g["VALOR_TOTAL"], g["_PCT"], int(g["N_ITENS"])]
        for col, val in zip(COLS, vals):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=10)
            c.alignment = _align()
            c.border = _border()
        ws[f"B{row}"].number_format = 'R$ #,##0.00'
        ws[f"C{row}"].number_format = "0.0%"
        ws.row_dimensions[row].height = 15
        row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = "TOTAL GERAL"
    ws[f"A{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"A{row}"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws[f"A{row}"].alignment = _align()
    ws[f"D{row}"].value = total_val
    ws[f"D{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"D{row}"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws[f"D{row}"].number_format = 'R$ #,##0.00'
    ws[f"D{row}"].alignment = _align()
    ws.row_dimensions[row].height = 20


# ── Estoque Morto ─────────────────────────────────────────────────────────────
def _gerar_estoque_morto(ws, df_estoque, df_saida, s_cod, s_data,
                         e_cod, e_grupo, e_desc, e_saldo, e_unid_e, e_vunit, meses_morto):
    MESES_CORTE = meses_morto

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 18

    ws.row_dimensions[1].height = 6

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"ESTOQUE MORTO — SEM MOVIMENTAÇÃO NOS ÚLTIMOS {MESES_CORTE} MESES"
    ws["A2"].fill = _fill(C_VERMELHO)
    ws["A2"].font = _font(bold=True, color=C_BRANCO, size=13)
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 10

    data_max  = df_saida[s_data].max()
    data_corte = data_max - pd.DateOffset(months=MESES_CORTE)
    cods_ativos = set(df_saida[df_saida[s_data] >= data_corte][s_cod].astype(str).unique())

    df_morto = df_estoque[
        (df_estoque[e_saldo] > 0) &
        (~df_estoque[e_cod].astype(str).isin(cods_ativos))
    ].copy()

    ws.merge_cells("A4:F4")
    if df_morto.empty:
        ws["A4"].value = "✔ Nenhum item sem movimentação encontrado no período!"
        ws["A4"].fill = _fill(C_VERDE_CLARO)
        ws["A4"].font = _font(bold=True, color=C_VERDE, size=11)
        ws["A4"].alignment = _align()
        return

    ws["A4"].value = (
        f"Período de referência: {data_corte.strftime('%d/%m/%Y')} a "
        f"{data_max.strftime('%d/%m/%Y')}  |  {len(df_morto)} itens parados"
    )
    ws["A4"].fill = _fill(C_VERMELHO_BG)
    ws["A4"].font = _font(size=10, color=C_VERMELHO)
    ws["A4"].alignment = _align()
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 10

    tem_valor = e_vunit and df_morto[e_vunit].sum() > 0
    if tem_valor:
        df_morto = df_morto.copy()
        df_morto["_VTOTAL"] = df_morto[e_saldo] * df_morto[e_vunit]

    hdrs = ["COD", "GRUPO", "DESCRIÇÃO", "UN", "SALDO", "VALOR PARADO (R$)"]
    COLS = ["A", "B", "C", "D", "E", "F"]
    for col, hdr in zip(COLS, hdrs):
        c = ws[f"{col}6"]
        c.value = hdr
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[6].height = 18

    row = 7
    for i, (_, item) in enumerate(df_morto.iterrows()):
        bg = C_CINZA if i % 2 == 0 else C_BRANCO
        vals = [
            item[e_cod] if e_cod else "",
            item[e_grupo] if e_grupo else "",
            item[e_desc] if e_desc else "",
            item[e_unid_e] if e_unid_e else "",
            item[e_saldo],
            item.get("_VTOTAL", "") if tem_valor else "",
        ]
        alinhamentos = ["center", "center", "left", "center", "center", "center"]
        for j, (col, val, al) in enumerate(zip(COLS, vals, alinhamentos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=9)
            c.alignment = _align(h=al, wrap=(j == 2))
            c.border = _border()
        if tem_valor and vals[5] != "":
            ws[f"F{row}"].number_format = 'R$ #,##0.00'
        ws.row_dimensions[row].height = 13
        row += 1


# ── Alerta de Estoque (Formatação Condicional) ────────────────────────────────
def _gerar_alerta_estoque(ws, df_estoque, df_saida, s_cod, s_quant, s_data,
                           e_cod, e_grupo, e_desc, e_saldo, e_unid_e, e_vunit):

    # Paleta específica do alerta
    C_RUPTURA_BG  = "FFD7D7"   # vermelho claro - sem estoque
    C_RUPTURA_FG  = "C00000"   # vermelho escuro
    C_CRITICO_BG  = "FFF2CC"   # amarelo claro - estoque crítico
    C_CRITICO_FG  = "7F6000"   # amarelo escuro
    C_OK_BG       = "E2EFDA"   # verde claro - ok
    C_OK_FG       = "375623"   # verde escuro

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    ws.row_dimensions[1].height = 6

    # Título
    ws.merge_cells("A2:H2")
    ws["A2"].value = "⚠  ALERTA DE ESTOQUE — CLASSIFICAÇÃO POR NÍVEL CRÍTICO"
    ws["A2"].fill = _fill(C_VERMELHO)
    ws["A2"].font = _font(bold=True, color=C_BRANCO, size=13)
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 30

    # Legenda
    ws.row_dimensions[3].height = 8
    legendas = [
        ("A4", "B4", "RUPTURA",  C_RUPTURA_FG, C_RUPTURA_BG, "Saldo = 0 (sem estoque)"),
        ("D4", "E4", "CRÍTICO",  C_CRITICO_FG, C_CRITICO_BG, "Saldo < consumo médio mensal"),
        ("F4", "H4", "NORMAL",   C_OK_FG,      C_OK_BG,      "Saldo adequado"),
    ]
    for col_label, col_desc, texto, fg, bg, desc in legendas:
        c = ws[col_label]
        c.value = texto
        c.fill = _fill(bg)
        c.font = _font(bold=True, color=fg, size=10)
        c.alignment = _align()
        c.border = _border()
        ws.merge_cells(f"{col_desc}4:{col_desc}4") if col_desc == col_label else None
        cd = ws[col_desc]
        cd.value = desc
        cd.fill = _fill(C_CINZA)
        cd.font = _font(size=9, color="555555")
        cd.alignment = _align(h="left")
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 8

    # ── Calcular consumo médio mensal por COD ────────────────────────────────
    df_s = df_saida.copy()
    df_s["_PERIODO"] = df_s[s_data].dt.to_period("M")
    n_meses = max(df_s["_PERIODO"].nunique(), 1)

    consumo_total = (
        df_s.groupby(s_cod)[s_quant]
        .sum()
        .reset_index()
        .rename(columns={s_quant: "_CONS_TOTAL"})
    )
    consumo_total["_CONS_MEDIO"] = consumo_total["_CONS_TOTAL"] / n_meses

    # Merge com estoque
    df = df_estoque.copy()
    df[e_cod] = df[e_cod].astype(str)
    consumo_total[s_cod] = consumo_total[s_cod].astype(str)
    df = df.merge(consumo_total[[s_cod, "_CONS_MEDIO"]], left_on=e_cod, right_on=s_cod, how="left")
    df["_CONS_MEDIO"] = df["_CONS_MEDIO"].fillna(0)

    # Calcular valor total se possível
    tem_valor = bool(e_vunit and df[e_vunit].sum() > 0)
    if tem_valor:
        df["_VT"] = df[e_saldo] * df[e_vunit]

    # Classificar
    def _classif(row):
        saldo = float(row[e_saldo]) if e_saldo else 0
        cons  = float(row["_CONS_MEDIO"])
        if saldo <= 0:
            return "RUPTURA"
        if cons > 0 and saldo < cons:
            return "CRÍTICO"
        return "NORMAL"

    df["_STATUS"] = df.apply(_classif, axis=1)

    n_ruptura = int((df["_STATUS"] == "RUPTURA").sum())
    n_critico = int((df["_STATUS"] == "CRÍTICO").sum())
    n_normal  = int((df["_STATUS"] == "NORMAL").sum())

    # ── Cards de resumo ──────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 16
    for col, label, val, fg, bg in [
        ("A", "RUPTURA",  n_ruptura, C_RUPTURA_FG, C_RUPTURA_BG),
        ("C", "CRÍTICO",  n_critico, C_CRITICO_FG, C_CRITICO_BG),
        ("E", "NORMAL",   n_normal,  C_OK_FG,      C_OK_BG),
        ("G", "TOTAL ITENS", len(df), C_BRANCO, C_AZUL_ESCURO),
    ]:
        ws[f"{col}6"].value = label
        ws[f"{col}6"].fill = _fill(bg)
        ws[f"{col}6"].font = _font(bold=True, color=fg, size=10)
        ws[f"{col}6"].alignment = _align()
        ws[f"{col}7"].value = val
        ws[f"{col}7"].fill = _fill(bg)
        ws[f"{col}7"].font = _font(bold=True, color=fg, size=20)
        ws[f"{col}7"].alignment = _align()
    ws.row_dimensions[7].height = 36
    ws.row_dimensions[8].height = 10

    # ── Cabeçalho da tabela ──────────────────────────────────────────────────
    hdrs = ["COD", "GRUPO", "DESCRIÇÃO", "UN", "SALDO ATUAL", "CONS. MÉDIO/MÊS", "VALOR EM ESTOQUE", "STATUS"]
    COLS = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for col, hdr in zip(COLS, hdrs):
        c = ws[f"{col}9"]
        c.value = hdr
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[9].height = 18

    # Ordenar: primeiro RUPTURA, depois CRÍTICO, depois NORMAL
    ordem = {"RUPTURA": 0, "CRÍTICO": 1, "NORMAL": 2}
    df = df.sort_values("_STATUS", key=lambda x: x.map(ordem)).reset_index(drop=True)

    COR_STATUS = {
        "RUPTURA": (C_RUPTURA_FG, C_RUPTURA_BG),
        "CRÍTICO": (C_CRITICO_FG, C_CRITICO_BG),
        "NORMAL":  (C_OK_FG,      C_OK_BG),
    }

    row = 10
    for _, item in df.iterrows():
        status = item["_STATUS"]
        fg, bg = COR_STATUS[status]

        cons_medio = item["_CONS_MEDIO"]
        valor = item.get("_VT", "") if tem_valor else ""

        vals = [
            item[e_cod] if e_cod else "",
            item[e_grupo] if e_grupo else "",
            item[e_desc] if e_desc else "",
            item[e_unid_e] if e_unid_e else "",
            float(item[e_saldo]) if e_saldo else 0,
            round(cons_medio, 1) if cons_medio > 0 else "—",
            valor,
            status,
        ]
        alinhamentos = ["center", "center", "left", "center", "center", "center", "center", "center"]

        for j, (col, val, al) in enumerate(zip(COLS, vals, alinhamentos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.border = _border()
            c.alignment = _align(h=al, wrap=(j == 2))

            if j == 7:  # coluna STATUS
                c.fill = _fill(bg)
                c.font = _font(bold=True, color=fg, size=9)
            elif status == "RUPTURA":
                c.fill = _fill(C_RUPTURA_BG)
                c.font = _font(bold=(j == 4), color=C_RUPTURA_FG if j == 4 else C_PRETO, size=9)
            elif status == "CRÍTICO":
                c.fill = _fill(C_CRITICO_BG)
                c.font = _font(bold=(j == 4), color=C_CRITICO_FG if j == 4 else C_PRETO, size=9)
            else:
                c.fill = _fill(C_BRANCO if row % 2 == 0 else C_CINZA)
                c.font = _font(size=9)

        if tem_valor and vals[6] != "":
            ws[f"G{row}"].number_format = "R$ #,##0.00"

        ws.row_dimensions[row].height = 13
        row += 1
