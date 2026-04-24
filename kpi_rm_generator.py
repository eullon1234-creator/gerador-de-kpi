"""
Gerador de KPI do RM - baseado na planilha de estoque do RM.

Input esperado (colunas):
    LOCESTOQUE | GRUPO | CODIGOPRD | PRODUTO | CODUN | SALDO | CUSTOMEDIO | VALORFINANCEIRO

Saída (5 abas):
    1. Resumo Executivo  — Dashboard com cards, curva ABC e Top 10 Grupos
    2. Estoque Completo  — Todo o estoque com classificação ABC
    3. Análise por Grupo — Estatísticas por grupo de material
    4. Top 50 Itens      — Os 50 itens com maior valor financeiro
    5. Estatísticas      — Indicadores estatísticos detalhados
"""
import warnings
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Paleta de Cores ───────────────────────────────────────────────────────────
C_AZUL_ESCURO  = "1F3864"
C_AZUL_MEDIO   = "2E74B5"
C_AZUL_CLARO   = "BDD7EE"
C_AZUL_SUAVE   = "DDEBF7"
C_LARANJA      = "ED7D31"
C_LARANJA_LEVE = "FCE4D6"
C_VERDE        = "375623"
C_VERDE_MEDIO  = "548235"
C_VERDE_CLARO  = "E2EFDA"
C_AMARELO      = "BF8F00"
C_AMARELO_BG   = "FFF2CC"
C_BRANCO       = "FFFFFF"
C_VERMELHO     = "C00000"
C_VERMELHO_BG  = "FCE4E4"
C_CINZA        = "F2F2F2"
C_CINZA_MEDIO  = "BFBFBF"
C_PRETO        = "000000"


# ── Helpers de Estilo ─────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_PRETO, size=10, name="Calibri", italic=False):
    return Font(bold=bold, color=color, size=size, name=name, italic=italic)


def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Utilitários de leitura ────────────────────────────────────────────────────
def _achar_col(colunas, *termos):
    """Encontra coluna por match exato, startswith ou contains (case-insensitive)."""
    colunas = list(colunas)
    for term in termos:
        for col in colunas:
            if str(col).upper().strip() == term.upper().strip():
                return col
    for term in termos:
        for col in colunas:
            if str(col).upper().strip().startswith(term.upper().strip()):
                return col
    for term in termos:
        for col in colunas:
            if term.upper().strip() in str(col).upper().strip():
                return col
    return None


def _normalizar_cols(df):
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.upper()
        .str.replace("\n", " ", regex=False)
        .str.replace(r"\s+", " ", regex=True)
    )
    return df


def _ler_planilha_rm(caminho):
    """Lê a primeira aba com dados da planilha de estoque do RM."""
    import openpyxl as _xl
    wb = _xl.load_workbook(caminho, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()

    if not sheets:
        raise ValueError("A planilha não possui nenhuma aba.")

    # Tenta encontrar uma aba com nome comum; se não, pega a primeira
    candidatos = ["ESTOQUE", "RM", "DADOS", "SHEET1", "PLAN1"]
    aba_escolhida = None
    for sh in sheets:
        if sh.strip().upper() in candidatos:
            aba_escolhida = sh
            break
    if aba_escolhida is None:
        aba_escolhida = sheets[0]

    # Detecta a linha de cabeçalho automaticamente (pode ter linhas de título)
    df_bruto = pd.read_excel(caminho, sheet_name=aba_escolhida, header=None)
    linha_cabecalho = 0
    for i in range(min(15, len(df_bruto))):
        linha = df_bruto.iloc[i]
        texto = " ".join(
            "" if pd.isna(valor) else str(valor).upper().strip()
            for valor in linha.tolist()
        )
        if ("LOCESTOQUE" in texto or "LOC ESTOQUE" in texto or "LOCAL" in texto) and \
           ("PRODUTO" in texto or "DESCRI" in texto) and \
           ("SALDO" in texto):
            linha_cabecalho = i
            break

    df = pd.read_excel(caminho, sheet_name=aba_escolhida, header=linha_cabecalho)
    df = _normalizar_cols(df)
    return df


# ── Função Principal ──────────────────────────────────────────────────────────
def gerar_kpi_rm(caminho_entrada, caminho_saida, top_grupos=10, top_itens=50,
                 limite_abc_a=0.80, limite_abc_b=0.95):
    """Gera o KPI do RM com 5 abas a partir da planilha de estoque."""

    df = _ler_planilha_rm(caminho_entrada)

    # Mapear colunas (aceita vários sinônimos)
    c_local  = _achar_col(df.columns, "LOCESTOQUE", "LOC ESTOQUE", "LOCAL ESTOQUE", "LOCAL")
    c_grupo  = _achar_col(df.columns, "GRUPO")
    c_cod    = _achar_col(df.columns, "CODIGOPRD", "CÓDIGO", "CODIGO", "COD PRD", "COD")
    c_prod   = _achar_col(df.columns, "PRODUTO", "DESCRIÇÃO", "DESCRICAO", "DESCRI")
    c_un     = _achar_col(df.columns, "CODUN", "UN", "UNIDADE")
    c_saldo  = _achar_col(df.columns, "SALDO", "QUANT")
    c_cmedio = _achar_col(df.columns, "CUSTOMEDIO", "CUSTO MEDIO", "CUSTO MÉDIO", "CUSTO")
    c_vfin   = _achar_col(df.columns, "VALORFINANCEIRO", "VALOR FINANCEIRO", "VALOR TOTAL", "VALOR")

    faltando = []
    for nome, col in [("GRUPO", c_grupo), ("CÓDIGO", c_cod), ("PRODUTO", c_prod),
                      ("SALDO", c_saldo), ("CUSTO MÉDIO", c_cmedio), ("VALOR FINANCEIRO", c_vfin)]:
        if col is None:
            faltando.append(nome)
    if faltando:
        raise ValueError(
            f"Colunas obrigatórias não encontradas: {', '.join(faltando)}. "
            f"Colunas detectadas: {list(df.columns)}"
        )

    # Limpar dados
    df[c_saldo]  = pd.to_numeric(df[c_saldo], errors="coerce").fillna(0)
    df[c_cmedio] = pd.to_numeric(df[c_cmedio], errors="coerce").fillna(0)
    df[c_vfin]   = pd.to_numeric(df[c_vfin], errors="coerce").fillna(0)
    df = df.dropna(subset=[c_cod]).copy()
    df[c_cod]   = df[c_cod].astype(str).str.strip()
    # Normaliza GRUPO: itens sem grupo ficam como "(SEM GRUPO)" para não
    # sumirem dos totais, mas são contados como 1 único "grupo especial"
    # (não aumenta artificialmente a contagem de grupos reais).
    df[c_grupo] = df[c_grupo].apply(
        lambda x: str(x).strip().upper() if pd.notna(x) and str(x).strip() else "(SEM GRUPO)"
    )
    df[c_prod]  = df[c_prod].astype(str).str.strip()

    # Mantém todos os itens (inclusive com saldo/valor zero) para bater com
    # o total de itens ativos da planilha original do RM.
    df_ativo = df.copy()

    # Identificar o Local de Estoque (pega o mais comum)
    if c_local and df_ativo[c_local].notna().any():
        local_est = str(df_ativo[c_local].mode().iloc[0]).strip()
    else:
        local_est = "ESTOQUE CENTRAL"

    data_hoje = datetime.now().strftime("%d/%m/%Y")

    # Classificação ABC (ordenada por valor financeiro desc)
    df_ativo = df_ativo.sort_values(c_vfin, ascending=False).reset_index(drop=True)
    total_valor = float(df_ativo[c_vfin].sum())
    df_ativo["_PCT"]  = df_ativo[c_vfin] / total_valor if total_valor > 0 else 0
    df_ativo["_ACUM"] = df_ativo["_PCT"].cumsum()

    def _classif(acum):
        if acum <= limite_abc_a:
            return "A"
        if acum <= limite_abc_b:
            return "B"
        return "C"

    df_ativo["_CLASSE"] = df_ativo["_ACUM"].apply(_classif)
    df_ativo["_RANK"]   = df_ativo.index + 1

    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Aba 1 — Resumo Executivo
    ws1 = wb.create_sheet("Resumo Executivo")
    _aba_resumo_executivo(ws1, df_ativo, c_grupo, c_cod, c_saldo, c_cmedio, c_vfin,
                          total_valor, local_est, data_hoje,
                          limite_abc_a, limite_abc_b, top_grupos)

    # Aba 2 — Estoque Completo
    ws2 = wb.create_sheet("Estoque Completo")
    _aba_estoque_completo(ws2, df_ativo, c_local, c_grupo, c_cod, c_prod, c_un,
                          c_saldo, c_cmedio, c_vfin, total_valor, local_est, data_hoje)

    # Aba 3 — Análise por Grupo
    ws3 = wb.create_sheet("Análise por Grupo")
    _aba_analise_grupo(ws3, df_ativo, c_grupo, c_cod, c_saldo, c_cmedio, c_vfin, total_valor)

    # Aba 4 — Top 50 Itens
    ws4 = wb.create_sheet("Top 50 Itens")
    _aba_top_itens(ws4, df_ativo, c_grupo, c_cod, c_prod, c_un, c_saldo, c_cmedio,
                   c_vfin, total_valor, top_itens)

    # Aba 5 — Estatísticas
    ws5 = wb.create_sheet("Estatísticas")
    _aba_estatisticas(ws5, df_ativo, c_grupo, c_saldo, c_cmedio, c_vfin,
                      total_valor, limite_abc_a, limite_abc_b)

    wb.save(caminho_saida)


# ─────────────────────────────────────────────────────────────────────────────
# ABA 1 — RESUMO EXECUTIVO
# ─────────────────────────────────────────────────────────────────────────────
def _aba_resumo_executivo(ws, df, c_grupo, c_cod, c_saldo, c_cmedio, c_vfin,
                          total_valor, local_est, data_hoje,
                          limite_abc_a, limite_abc_b, top_grupos):

    larguras = {"A": 3, "B": 18, "C": 18, "D": 18, "E": 16, "F": 16, "G": 8,
                "H": 32, "I": 12, "J": 16, "K": 10, "L": 10}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 8

    # ── Título Principal ────────────────────────────────────────────────────
    ws.merge_cells("B2:L2")
    c = ws["B2"]
    c.value = "DASHBOARD KPI — GESTÃO DE ESTOQUE"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=18)
    c.alignment = _align()
    ws.row_dimensions[2].height = 36

    # Subtítulo
    ws.merge_cells("B3:L3")
    c = ws["B3"]
    c.value = f"{local_est}  |  Data: {data_hoje}"
    c.fill = _fill(C_AZUL_SUAVE)
    c.font = _font(bold=True, color=C_AZUL_ESCURO, size=11, italic=True)
    c.alignment = _align()
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 8

    # ── KPIs — Cards superiores ────────────────────────────────────────────
    total_itens     = len(df)
    # Conta grupos reais (ignora "(SEM GRUPO)" e nulos)
    grupos_reais = df[c_grupo][df[c_grupo].notna() & (df[c_grupo] != "(SEM GRUPO)")]
    total_grupos    = grupos_reais.nunique()
    saldo_total     = float(df[c_saldo].sum())
    custo_medio_ger = df[c_cmedio].mean() if len(df) else 0

    mask_epi = df[c_grupo].str.contains("EPI", na=False) | df[c_grupo].str.contains("EPC", na=False)
    valor_epi = df.loc[mask_epi, c_vfin].sum()
    pct_epi   = (valor_epi / total_valor) if total_valor > 0 else 0

    cards = [
        ("B", "D", "VALOR TOTAL\nDO ESTOQUE",    total_valor,     "moeda",  C_AZUL_MEDIO),
        ("E", "E", "TOTAL DE\nITENS ATIVOS",    total_itens,     "int",    C_AZUL_MEDIO),
        ("F", "F", "GRUPOS DE\nMATERIAIS",      total_grupos,    "int",    C_AZUL_MEDIO),
        ("G", "I", "QUANTIDADE\nTOTAL (SALDO)", saldo_total,     "int",    C_AZUL_MEDIO),
        ("J", "J", "CUSTO MÉDIO\nPOR ITEM",     custo_medio_ger, "moeda",  C_AZUL_MEDIO),
        ("K", "L", "% EPI/EPC\nNO ESTOQUE",     pct_epi,         "pct",    C_VERDE_MEDIO),
    ]

    # Cabeçalho dos cards
    for col_i, col_f, label, _, _, cor in cards:
        if col_i != col_f:
            ws.merge_cells(f"{col_i}5:{col_f}5")
        c = ws[f"{col_i}5"]
        c.value = label
        c.fill = _fill(cor)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align(wrap=True)
        c.border = _border(C_BRANCO)
    ws.row_dimensions[5].height = 42

    # Valor dos cards
    for col_i, col_f, _, valor, tipo, cor in cards:
        if col_i != col_f:
            ws.merge_cells(f"{col_i}6:{col_f}6")
        c = ws[f"{col_i}6"]
        c.value = valor
        if tipo == "moeda":
            c.number_format = 'R$ #,##0.00'
        elif tipo == "int":
            c.number_format = '#,##0'
        elif tipo == "pct":
            c.number_format = '0.0%'
        c.fill = _fill(C_BRANCO)
        c.font = _font(bold=True, color=cor, size=16)
        c.alignment = _align()
        c.border = _border(cor)
    ws.row_dimensions[6].height = 46

    ws.row_dimensions[7].height = 12

    # ── Curva ABC + Top 10 Grupos (lado a lado) ─────────────────────────────
    # Título CURVA ABC (cols B-F)
    ws.merge_cells("B9:F9")
    c = ws["B9"]
    c.value = "CURVA ABC — CLASSIFICAÇÃO POR VALOR FINANCEIRO"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()

    # Título TOP 10 GRUPOS (cols G-L)
    ws.merge_cells("G9:L9")
    c = ws["G9"]
    c.value = f"TOP {top_grupos} GRUPOS POR VALOR FINANCEIRO"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()
    ws.row_dimensions[9].height = 22

    # Cabeçalho Curva ABC
    hdrs_abc = ["CLASSE", "CRITÉRIO", "QTD ITENS", "% ITENS", "VALOR (R$)", "% VALOR"]
    cols_abc = ["B", "C", "D", "E", "F"]
    # Precisa de 6 colunas - vou mesclar B-F corretamente
    cols_abc = ["B", "C", "D", "E", "F"]
    headers_abc = [
        ("B10", "CLASSE"),
        ("C10", "CRITÉRIO"),
        ("D10", "QTD ITENS"),
        ("E10", "% ITENS"),
        ("F10", "VALOR (R$) / % VALOR"),
    ]

    # Layout: B=CLASSE, C=CRITÉRIO, D=QTD, E=%ITENS, F=VALOR — % VALOR fica junto
    # Melhor: B CLASSE, C CRITÉRIO (merge largo), D QTD ITENS, E %ITENS, F VALOR, G %VALOR
    # Mas G é ocupado pelo Top 10. Então compactamos em 5 colunas

    for celula, texto in headers_abc:
        c = ws[celula]
        c.value = texto
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[10].height = 28

    # Dados Curva ABC
    classes_info = []
    for cls in ["A", "B", "C"]:
        sub = df[df["_CLASSE"] == cls]
        classes_info.append({
            "classe": cls,
            "qtd": len(sub),
            "pct_itens": len(sub) / len(df) if len(df) else 0,
            "valor": float(sub[c_vfin].sum()),
            "pct_valor": float(sub[c_vfin].sum()) / total_valor if total_valor > 0 else 0,
        })

    criterios = [
        f"0% — {int(limite_abc_a*100)}% do valor",
        f"{int(limite_abc_a*100)}% — {int(limite_abc_b*100)}% do valor",
        f"{int(limite_abc_b*100)}% — 100% do valor",
    ]
    cores_classe = [
        ("375623", "E2EFDA"),  # A — verde
        ("9C5700", "FFF2CC"),  # B — amarelo/laranja
        ("9C0006", "FFC7CE"),  # C — vermelho
    ]

    row = 11
    for i, info in enumerate(classes_info):
        fg, bg = cores_classe[i]

        # CLASSE (B)
        c = ws[f"B{row}"]
        c.value = info["classe"]
        c.fill = _fill(fg)
        c.font = _font(bold=True, color=C_BRANCO, size=14)
        c.alignment = _align()
        c.border = _border()

        # CRITÉRIO (C)
        c = ws[f"C{row}"]
        c.value = criterios[i]
        c.fill = _fill(bg)
        c.font = _font(size=9)
        c.alignment = _align(h="left")
        c.border = _border()

        # QTD ITENS (D)
        c = ws[f"D{row}"]
        c.value = info["qtd"]
        c.number_format = '#,##0'
        c.fill = _fill(bg)
        c.font = _font(bold=True, size=10)
        c.alignment = _align()
        c.border = _border()

        # % ITENS (E)
        c = ws[f"E{row}"]
        c.value = info["pct_itens"]
        c.number_format = '0.0%'
        c.fill = _fill(bg)
        c.font = _font(size=10)
        c.alignment = _align()
        c.border = _border()

        # VALOR (F) — com % abaixo visualmente
        c = ws[f"F{row}"]
        c.value = info["valor"]
        c.number_format = 'R$ #,##0.00'
        c.fill = _fill(bg)
        c.font = _font(bold=True, size=10, color=fg)
        c.alignment = _align()
        c.border = _border()

        ws.row_dimensions[row].height = 24
        row += 1

    # Linha adicional com % do valor (em baixo da curva)
    c = ws[f"B{row}"]
    c.value = "TOTAL"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.alignment = _align()
    c.border = _border()

    ws.merge_cells(f"C{row}:C{row}")
    c = ws[f"C{row}"]
    c.value = "100% do estoque"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=9)
    c.alignment = _align(h="left")
    c.border = _border()

    c = ws[f"D{row}"]
    c.value = len(df)
    c.number_format = '#,##0'
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.alignment = _align()
    c.border = _border()

    c = ws[f"E{row}"]
    c.value = 1.0
    c.number_format = '0.0%'
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.alignment = _align()
    c.border = _border()

    c = ws[f"F{row}"]
    c.value = total_valor
    c.number_format = 'R$ #,##0.00'
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.alignment = _align()
    c.border = _border()

    ws.row_dimensions[row].height = 22

    # ── Top 10 Grupos (direita: cols G a L) ─────────────────────────────────
    por_grupo = (
        df.groupby(c_grupo)
        .agg(QTD=(c_cod, "count"), VALOR=(c_vfin, "sum"))
        .reset_index()
        .sort_values("VALOR", ascending=False)
        .head(top_grupos)
        .reset_index(drop=True)
    )
    por_grupo["_PCT"]  = por_grupo["VALOR"] / total_valor if total_valor > 0 else 0
    por_grupo["_ACUM"] = por_grupo["_PCT"].cumsum()

    hdrs_g = ["RANK", "GRUPO", "QTD ITENS", "VALOR (R$)", "% TOTAL", "% ACUM."]
    cols_g = ["G", "H", "I", "J", "K", "L"]
    for col, hdr in zip(cols_g, hdrs_g):
        c = ws[f"{col}10"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align(wrap=True)
        c.border = _border()

    row_g = 11
    for i, (_, g) in enumerate(por_grupo.iterrows(), start=1):
        bg = C_BRANCO if i % 2 == 1 else C_CINZA
        valores = [i, g[c_grupo], int(g["QTD"]), g["VALOR"], g["_PCT"], g["_ACUM"]]
        alinhamentos = ["center", "left", "center", "right", "center", "center"]

        for j, (col, val, al) in enumerate(zip(cols_g, valores, alinhamentos)):
            c = ws[f"{col}{row_g}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=9, bold=(j == 0))
            c.alignment = _align(h=al, wrap=(j == 1))
            c.border = _border()
            if j == 3:
                c.number_format = 'R$ #,##0.00'
            elif j == 4 or j == 5:
                c.number_format = '0.00%'
        ws.row_dimensions[row_g].height = 20
        row_g += 1


# ─────────────────────────────────────────────────────────────────────────────
# ABA 2 — ESTOQUE COMPLETO
# ─────────────────────────────────────────────────────────────────────────────
def _aba_estoque_completo(ws, df, c_local, c_grupo, c_cod, c_prod, c_un,
                          c_saldo, c_cmedio, c_vfin, total_valor, local_est, data_hoje):

    larguras = {
        "A": 6, "B": 22, "C": 28, "D": 18, "E": 52, "F": 7,
        "G": 11, "H": 15, "I": 18, "J": 10, "K": 10, "L": 11, "M": 18
    }
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 8

    # Título
    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = f"ESTOQUE COMPLETO COM CLASSIFICAÇÃO ABC — {local_est} — {data_hoje}"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=13)
    c.alignment = _align()
    ws.row_dimensions[2].height = 32

    # Cabeçalho
    hdrs = ["RANK", "LOCAL ESTOQUE", "GRUPO", "CÓDIGO", "PRODUTO", "UN",
            "SALDO", "CUSTO MÉDIO (R$)", "VALOR FINANCEIRO (R$)",
            "% DO TOTAL", "% ACUM.", "CLASSE ABC", "RANKING ACUMULADO"]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    for col, hdr in zip(cols, hdrs):
        c = ws[f"{col}3"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[3].height = 32

    cores_classe = {
        "A": ("375623", "E2EFDA"),
        "B": ("9C5700", "FFF2CC"),
        "C": ("9C0006", "FFC7CE"),
    }

    row = 4
    acumulado_valor = 0.0
    for i, (_, item) in enumerate(df.iterrows(), start=1):
        classe = item["_CLASSE"]
        fg, bg_cls = cores_classe[classe]
        bg = C_BRANCO if i % 2 == 1 else C_CINZA
        acumulado_valor += float(item[c_vfin])

        valores = [
            i,
            item[c_local] if c_local else local_est,
            item[c_grupo],
            item[c_cod],
            item[c_prod],
            item[c_un] if c_un else "",
            float(item[c_saldo]),
            float(item[c_cmedio]),
            float(item[c_vfin]),
            item["_PCT"],
            item["_ACUM"],
            classe,
            acumulado_valor,
        ]
        alinhamentos = ["center", "center", "center", "center", "left", "center",
                        "right", "right", "right", "center", "center", "center", "right"]
        formatos = [None, None, None, None, None, None,
                    '#,##0.00', 'R$ #,##0.00', 'R$ #,##0.00',
                    '0.00%', '0.00%', None, 'R$ #,##0.00']

        for j, (col, val, al, fmt) in enumerate(zip(cols, valores, alinhamentos, formatos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.alignment = _align(h=al, wrap=(j == 4))
            c.border = _border()
            if fmt:
                c.number_format = fmt
            if j == 11:  # CLASSE ABC
                c.fill = _fill(fg)
                c.font = _font(bold=True, color=C_BRANCO, size=10)
            else:
                c.fill = _fill(bg)
                c.font = _font(size=9)

        ws.row_dimensions[row].height = 14
        row += 1

    # Freeze panes
    ws.freeze_panes = "A4"
    # Auto-filter
    ws.auto_filter.ref = f"A3:M{row-1}"


# ─────────────────────────────────────────────────────────────────────────────
# ABA 3 — ANÁLISE POR GRUPO
# ─────────────────────────────────────────────────────────────────────────────
def _aba_analise_grupo(ws, df, c_grupo, c_cod, c_saldo, c_cmedio, c_vfin, total_valor):

    larguras = {"A": 6, "B": 42, "C": 10, "D": 12, "E": 14, "F": 14, "G": 14,
                "H": 18, "I": 10, "J": 12}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 8

    # Título
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "ANÁLISE COMPLETA POR GRUPO DE MATERIAL"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=14)
    c.alignment = _align()
    ws.row_dimensions[2].height = 32

    # Cabeçalho
    hdrs = ["RANK", "GRUPO", "QTD ITENS", "SALDO TOTAL", "CUSTO MÉDIO",
            "CUSTO MÁXIMO", "CUSTO MÍNIMO", "VALOR TOTAL (R$)",
            "% DO TOTAL", "% ACUMULADO"]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    for col, hdr in zip(cols, hdrs):
        c = ws[f"{col}3"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[3].height = 32

    # Agrega por grupo
    agg = (
        df.groupby(c_grupo)
        .agg(
            QTD_ITENS=(c_cod, "count"),
            SALDO_TOTAL=(c_saldo, "sum"),
            CUSTO_MEDIO=(c_cmedio, "mean"),
            CUSTO_MAX=(c_cmedio, "max"),
            CUSTO_MIN=(c_cmedio, "min"),
            VALOR_TOTAL=(c_vfin, "sum"),
        )
        .reset_index()
        .sort_values("VALOR_TOTAL", ascending=False)
        .reset_index(drop=True)
    )
    agg["_PCT"] = agg["VALOR_TOTAL"] / total_valor if total_valor > 0 else 0
    agg["_ACUM"] = agg["_PCT"].cumsum()

    row = 4
    for i, (_, g) in enumerate(agg.iterrows(), start=1):
        bg = C_BRANCO if i % 2 == 1 else C_CINZA

        valores = [
            i,
            g[c_grupo],
            int(g["QTD_ITENS"]),
            float(g["SALDO_TOTAL"]),
            float(g["CUSTO_MEDIO"]),
            float(g["CUSTO_MAX"]),
            float(g["CUSTO_MIN"]),
            float(g["VALOR_TOTAL"]),
            g["_PCT"],
            g["_ACUM"],
        ]
        alinhamentos = ["center", "left", "center", "right", "right", "right",
                        "right", "right", "center", "center"]
        formatos = [None, None, '#,##0', '#,##0.00', 'R$ #,##0.00', 'R$ #,##0.00',
                    'R$ #,##0.00', 'R$ #,##0.00', '0.00%', '0.00%']

        for j, (col, val, al, fmt) in enumerate(zip(cols, valores, alinhamentos, formatos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=9, bold=(j == 0))
            c.alignment = _align(h=al, wrap=(j == 1))
            c.border = _border()
            if fmt:
                c.number_format = fmt

        ws.row_dimensions[row].height = 16
        row += 1

    # Total geral
    ws[f"A{row}"].value = "TOTAL"
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"A{row}"].font = _font(bold=True, color=C_BRANCO, size=10)
    ws[f"A{row}"].alignment = _align()
    ws[f"A{row}"].border = _border()

    total_cells = [
        ("C", len(df), '#,##0'),
        ("D", float(df[c_saldo].sum()), '#,##0.00'),
        ("E", "", None),
        ("F", "", None),
        ("G", "", None),
        ("H", total_valor, 'R$ #,##0.00'),
        ("I", 1.0, '0.00%'),
        ("J", "", None),
    ]
    for col, val, fmt in total_cells:
        c = ws[f"{col}{row}"]
        c.value = val
        c.fill = _fill(C_AZUL_ESCURO)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.alignment = _align()
        c.border = _border()
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{row-1}"


# ─────────────────────────────────────────────────────────────────────────────
# ABA 4 — TOP 50 ITENS
# ─────────────────────────────────────────────────────────────────────────────
def _aba_top_itens(ws, df, c_grupo, c_cod, c_prod, c_un, c_saldo, c_cmedio,
                   c_vfin, total_valor, top_itens):

    larguras = {"A": 6, "B": 32, "C": 18, "D": 54, "E": 7, "F": 11,
                "G": 15, "H": 20, "I": 10}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 8

    # Título
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"TOP {top_itens} ITENS — MAIOR VALOR FINANCEIRO NO ESTOQUE"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=14)
    c.alignment = _align()
    ws.row_dimensions[2].height = 32

    # Cabeçalho
    hdrs = ["RANK", "GRUPO", "CÓDIGO", "PRODUTO", "UN", "SALDO",
            "CUSTO MÉDIO (R$)", "VALOR FINANCEIRO (R$)", "% DO TOTAL"]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for col, hdr in zip(cols, hdrs):
        c = ws[f"{col}3"]
        c.value = hdr
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=9)
        c.alignment = _align(wrap=True)
        c.border = _border()
    ws.row_dimensions[3].height = 30

    df_top = df.head(top_itens)

    # Destaca as 3 primeiras posições com cores especiais
    cores_top = [
        ("BF8F00", "FFE699"),  # 1º - ouro
        ("808080", "E7E6E6"),  # 2º - prata
        ("A6665A", "F4E7E1"),  # 3º - bronze
    ]

    row = 4
    for i, (_, item) in enumerate(df_top.iterrows(), start=1):
        if i <= 3:
            fg, bg = cores_top[i - 1]
            font_bold = True
        else:
            fg = C_PRETO
            bg = C_BRANCO if i % 2 == 1 else C_CINZA
            font_bold = False

        valores = [
            i,
            item[c_grupo],
            item[c_cod],
            item[c_prod],
            item[c_un] if c_un else "",
            float(item[c_saldo]),
            float(item[c_cmedio]),
            float(item[c_vfin]),
            item["_PCT"],
        ]
        alinhamentos = ["center", "left", "center", "left", "center",
                        "right", "right", "right", "center"]
        formatos = [None, None, None, None, None, '#,##0.00',
                    'R$ #,##0.00', 'R$ #,##0.00', '0.00%']

        for j, (col, val, al, fmt) in enumerate(zip(cols, valores, alinhamentos, formatos)):
            c = ws[f"{col}{row}"]
            c.value = val
            c.fill = _fill(bg)
            c.font = _font(size=10, bold=(font_bold or j == 0), color=fg if font_bold else C_PRETO)
            c.alignment = _align(h=al, wrap=(j == 1 or j == 3))
            c.border = _border()
            if fmt:
                c.number_format = fmt

        ws.row_dimensions[row].height = 18 if i <= 3 else 15
        row += 1

    # Linha total do TOP
    ws[f"A{row}"].value = f"TOTAL TOP {top_itens}"
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].fill = _fill(C_AZUL_ESCURO)
    ws[f"A{row}"].font = _font(bold=True, color=C_BRANCO, size=11)
    ws[f"A{row}"].alignment = _align()
    ws[f"A{row}"].border = _border()

    total_top = float(df_top[c_vfin].sum())
    pct_top   = total_top / total_valor if total_valor > 0 else 0

    c = ws[f"H{row}"]
    c.value = total_top
    c.number_format = 'R$ #,##0.00'
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()
    c.border = _border()

    c = ws[f"I{row}"]
    c.value = pct_top
    c.number_format = '0.00%'
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.alignment = _align()
    c.border = _border()

    ws.row_dimensions[row].height = 26

    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────────────────────
# ABA 5 — ESTATÍSTICAS
# ─────────────────────────────────────────────────────────────────────────────
def _aba_estatisticas(ws, df, c_grupo, c_saldo, c_cmedio, c_vfin,
                      total_valor, limite_abc_a, limite_abc_b):

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    ws.row_dimensions[1].height = 6

    # Título
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "ESTATÍSTICAS ANALÍTICAS DO ESTOQUE"
    c.fill = _fill(C_AZUL_ESCURO)
    c.font = _font(bold=True, color=C_BRANCO, size=14)
    c.alignment = _align()
    ws.row_dimensions[2].height = 32

    def secao(titulo, linha_inicio):
        ws.merge_cells(f"A{linha_inicio}:B{linha_inicio}")
        c = ws[f"A{linha_inicio}"]
        c.value = titulo
        c.fill = _fill(C_AZUL_MEDIO)
        c.font = _font(bold=True, color=C_BRANCO, size=11)
        c.alignment = _align(h="left")
        c.border = _border()
        ws.row_dimensions[linha_inicio].height = 22

    def linha(label, valor, row, fmt=None, bold_valor=False):
        c = ws[f"A{row}"]
        c.value = label
        c.fill = _fill(C_AZUL_SUAVE)
        c.font = _font(size=10)
        c.alignment = _align(h="left")
        c.border = _border()

        c = ws[f"B{row}"]
        c.value = valor
        c.fill = _fill(C_BRANCO)
        c.font = _font(bold=bold_valor, size=10, color=C_AZUL_ESCURO)
        c.alignment = _align(h="right")
        c.border = _border()
        if fmt:
            c.number_format = fmt
        ws.row_dimensions[row].height = 18

    # ── Indicadores Gerais ──
    row = 3
    secao("INDICADORES GERAIS", row); row += 1
    linha("Total de itens no estoque",       len(df),                      row, '#,##0'); row += 1
    linha("Total de grupos de materiais",    int(df[c_grupo].nunique()),   row, '#,##0'); row += 1
    linha("Valor total do estoque",          total_valor,                  row, 'R$ #,##0.00', bold_valor=True); row += 1
    linha("Saldo total (quantidade)",        float(df[c_saldo].sum()),     row, '#,##0.00'); row += 1
    linha("Média de saldo por item",         float(df[c_saldo].mean()) if len(df) else 0, row, '#,##0.00'); row += 1

    row += 1

    # ── Custo Médio ──
    secao("CUSTO MÉDIO", row); row += 1
    linha("Custo médio geral dos itens",     float(df[c_cmedio].mean()) if len(df) else 0,   row, 'R$ #,##0.00'); row += 1
    linha("Mediana do custo médio",          float(df[c_cmedio].median()) if len(df) else 0, row, 'R$ #,##0.00'); row += 1
    linha("Custo médio mínimo",              float(df[c_cmedio].min()) if len(df) else 0,    row, 'R$ #,##0.00'); row += 1
    linha("Custo médio máximo",              float(df[c_cmedio].max()) if len(df) else 0,    row, 'R$ #,##0.00'); row += 1
    linha("Desvio padrão do custo médio",    float(df[c_cmedio].std()) if len(df) > 1 else 0, row, 'R$ #,##0.00'); row += 1

    row += 1

    # ── Valor Financeiro por Item ──
    secao("VALOR FINANCEIRO POR ITEM", row); row += 1
    linha("Menor valor financeiro",          float(df[c_vfin].min()) if len(df) else 0,    row, 'R$ #,##0.00'); row += 1
    linha("Maior valor financeiro",          float(df[c_vfin].max()) if len(df) else 0,    row, 'R$ #,##0.00'); row += 1
    linha("Valor médio por item",            float(df[c_vfin].mean()) if len(df) else 0,   row, 'R$ #,##0.00'); row += 1
    linha("Mediana do valor financeiro",     float(df[c_vfin].median()) if len(df) else 0, row, 'R$ #,##0.00'); row += 1
    linha("Desvio padrão do valor",          float(df[c_vfin].std()) if len(df) > 1 else 0, row, 'R$ #,##0.00'); row += 1

    row += 1

    # ── Curva ABC ──
    secao("CURVA ABC", row); row += 1
    for cls in ["A", "B", "C"]:
        sub = df[df["_CLASSE"] == cls]
        qtd = len(sub)
        pct = qtd / len(df) if len(df) else 0
        val = float(sub[c_vfin].sum())
        linha(f"Classe {cls} — itens (qtd)",   qtd, row, '#,##0'); row += 1
        linha(f"Classe {cls} — itens (%)",     pct, row, '0.00%'); row += 1
        linha(f"Classe {cls} — valor total",   val, row, 'R$ #,##0.00', bold_valor=True); row += 1

    row += 1

    # ── Concentração / Análise adicional ──
    secao("ANÁLISE DE CONCENTRAÇÃO", row); row += 1

    # Top 10 itens % do valor
    top10_valor = float(df.head(10)[c_vfin].sum())
    pct_top10   = top10_valor / total_valor if total_valor > 0 else 0
    linha("Top 10 itens — valor",        top10_valor, row, 'R$ #,##0.00'); row += 1
    linha("Top 10 itens — % do total",   pct_top10,   row, '0.00%'); row += 1

    top50_valor = float(df.head(50)[c_vfin].sum())
    pct_top50   = top50_valor / total_valor if total_valor > 0 else 0
    linha("Top 50 itens — valor",        top50_valor, row, 'R$ #,##0.00'); row += 1
    linha("Top 50 itens — % do total",   pct_top50,   row, '0.00%'); row += 1

    # Grupo líder
    grupo_lider = df.groupby(c_grupo)[c_vfin].sum().idxmax()
    valor_lider = df.groupby(c_grupo)[c_vfin].sum().max()
    linha("Grupo com maior valor",       str(grupo_lider),             row); row += 1
    linha("Valor do grupo líder",        float(valor_lider),           row, 'R$ #,##0.00'); row += 1
    linha("% do grupo líder no total",   float(valor_lider)/total_valor if total_valor else 0, row, '0.00%'); row += 1
